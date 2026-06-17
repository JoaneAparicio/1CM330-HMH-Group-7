"""
compare_stats.py
================
Reads MH and Memetic run-level fitness values from Excel result files,
produces box plots per case, and runs Wilcoxon signed-rank tests
(or paired t-tests as a fallback when sample variance is zero).

HOW TO USE
----------
1. Run your memetic algorithm on the Table 8 sub-instances and produce
   a results_memetic_table8.xlsx in the same format as the base file.
2. Set the four path constants in CONFIGURATION below.
3. Run:  python compare_stats.py

OUTPUT
------
- compare_base_cases.png   : box plots for the 4 base instances
- compare_table8.png       : box plots for the 7 Table-8 sub-instances
- statistical_tests.xlsx   : full test results table
"""

from __future__ import annotations
import warnings
from pathlib import Path

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# ✏️  CONFIGURATION – set your file paths here
# ============================================================================
RESULTS_DIR = Path("output")  # directory where your result Excel files are located

MH_BASE_XLSX      = RESULTS_DIR / "results_mh_base.xlsx"
MH_TABLE8_XLSX    = RESULTS_DIR / "results_mh_table8.xlsx"
MEMETIC_BASE_XLSX = RESULTS_DIR / "results_memetic_base.xlsx"       # ← your memetic BASE results
MEMETIC_T8_XLSX   = RESULTS_DIR / "results_memetic_table8.xlsx"    # ← your memetic TABLE 8 results

OUTPUT_DIR = Path("output")

# ============================================================================
# HELPERS – Excel reading
# ============================================================================

def _read_run_fitness(xlsx_path: str | Path) -> dict[str, list[float]]:
    """
    Parse a results Excel file (MH or Memetic format) and return
    {instance_label: [run1_fitness, run2_fitness, ...]} for every detail sheet.

    Works for both file formats:
      - detail sheets have a "Run | Fitness | ..." table
      - the run rows start after the header row that contains the word "Run"
        and end at the first blank or "Statistic" row.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data: dict[str, list[float]] = {}

    for sname in wb.sheetnames:
        if sname.lower() == "summary":
            continue
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))

        # find the header row that contains "Run"
        header_row_idx = None
        for i, row in enumerate(rows):
            vals = [str(v).strip() if v is not None else "" for v in row]
            if "Run" in vals:
                header_row_idx = i
                break
        if header_row_idx is None:
            continue

        # fitness is in column index 1 (second column, 0-based)
        fitnesses: list[float] = []
        for row in rows[header_row_idx + 1:]:
            if row[0] is None or str(row[0]).strip() in ("", "Statistic"):
                break
            try:
                run_no = int(row[0])  # noqa: F841 – just checking it's a run row
                fitnesses.append(float(row[1]))
            except (TypeError, ValueError):
                break

        if fitnesses:
            # use a cleaned version of the sheet name as key
            label = sname.replace("__", " ").replace("_", " ").strip()
            data[label] = fitnesses

    return data


def _read_summary_ph(xlsx_path: str | Path) -> dict[str, float]:
    """Return {instance_label: ph_obj} from the Summary sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Summary"]
    rows = list(ws.iter_rows(values_only=True))

    ph: dict[str, float] = {}
    header_idx = None
    for i, row in enumerate(rows):
        vals = [str(v).strip() if v is not None else "" for v in row]
        if "Instance" in vals:
            header_idx = i
            break
    if header_idx is None:
        return ph

    # find column indices
    h = [str(v).strip() if v is not None else "" for v in rows[header_idx]]
    try:
        inst_col = h.index("Instance")
        ph_col   = next(i for i, x in enumerate(h) if "PH" in x)
    except (ValueError, StopIteration):
        return ph

    for row in rows[header_idx + 1:]:
        if row[inst_col] is None:
            break
        try:
            ph[str(row[inst_col]).strip()] = float(row[ph_col])
        except (TypeError, ValueError):
            pass
    return ph


# ============================================================================
# STATISTICAL TEST
# ============================================================================

def _conclusion(p: float, mean_a: float, mean_b: float) -> str:
    """MH=a, Memetic=b. Lower is better."""
    if p < 0.05:
        return "MH better" if mean_a < mean_b else "Memetic better"
    return "no sig. diff."


def _compute_arpd(runs: list[float], bks: float) -> float:
    """ARPD = mean((f_i - BKS) / BKS * 100) over all runs."""
    if not runs or bks <= 0:
        return np.nan
    arr = np.array(runs, dtype=float)
    return float(np.mean((arr - bks) / bks * 100))


def _run_test(a: list[float], b: list[float]) -> dict:
    """
    Always runs BOTH Wilcoxon signed-rank and paired t-test.
    Returns stats for both, plus a combined conclusion.
    If differences have zero variance, both tests are marked N/A.
    """
    a, b = np.array(a, dtype=float), np.array(b, dtype=float)
    diffs = a - b
    mean_a, mean_b = float(np.mean(a)), float(np.mean(b))

    # Trivial case: all differences identical
    if np.std(diffs) < 1e-9:
        if np.abs(np.mean(diffs)) < 1e-9:
            conc = "identical"
        else:
            conc = "MH better" if mean_a < mean_b else "Memetic better"
        return {
            "w_stat": np.nan, "w_p": np.nan,
            "t_stat": np.nan, "t_p": np.nan,
            "conclusion": conc,
        }

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")

        # Wilcoxon
        try:
            w_stat, w_p = stats.wilcoxon(a, b, alternative="two-sided")
        except ValueError:
            w_stat, w_p = np.nan, np.nan

        # Paired t-test
        try:
            t_stat, t_p = stats.ttest_rel(a, b)
        except Exception:
            t_stat, t_p = np.nan, np.nan

    # Conclusion: both tests must agree (p<0.05) to claim significance
    w_sig = (not np.isnan(w_p)) and w_p < 0.05
    t_sig = (not np.isnan(t_p)) and t_p < 0.05
    if w_sig and t_sig:
        conc = "MH better" if mean_a < mean_b else "Memetic better"
    elif w_sig or t_sig:
        conc = f"{'MH' if mean_a < mean_b else 'Memetic'} better (1/2 tests)"
    else:
        conc = "no sig. diff."

    return {
        "w_stat": float(w_stat), "w_p": float(w_p),
        "t_stat": float(t_stat), "t_p": float(t_p),
        "conclusion": conc,
    }


# ============================================================================
# BOX PLOT
# ============================================================================

COLOR_MH      = "#2E75B6"
COLOR_MEMETIC = "#70AD47"
COLOR_PH      = "#C00000"
FACE_ALPHA    = 0.80


def _draw_box(ax, pos, data, color):
    """Draw a box plot with individual run points (strip chart) overlaid."""
    arr = np.array(data, dtype=float)
    q1, q3 = float(np.percentile(arr, 25)), float(np.percentile(arr, 75))
    flat = (q3 - q1) < 1e-6  # degenerate box → make points more prominent

    bp = ax.boxplot(
        arr,
        positions=[pos],
        widths=0.45,
        patch_artist=True,
        notch=False,
        medianprops=dict(color="white", linewidth=2.5, zorder=5),
        boxprops=dict(facecolor=color, alpha=FACE_ALPHA, linewidth=1.2,
                      edgecolor="white"),
        whiskerprops=dict(color=color, linewidth=1.5, linestyle="-"),
        capprops=dict(color=color, linewidth=2.0),
        showfliers=False,  # outliers shown via scatter below
        zorder=4,
    )

    # Strip chart: one dot per run, with fixed-seed jitter
    rng = np.random.default_rng(42)
    jitter = rng.uniform(-0.14, 0.14, size=len(arr))
    pt_size  = 28 if flat else 18
    pt_alpha = 0.90 if flat else 0.65
    ax.scatter(
        pos + jitter, arr,
        s=pt_size, color=color, alpha=pt_alpha,
        edgecolors="white", linewidths=0.6,
        zorder=6,
    )
    # annotate mean and best
    mean_v = float(np.mean(arr))
    best_v = float(np.min(arr))
    ax.annotate(f"μ={mean_v:.1f}", xy=(pos, mean_v),
                xytext=(pos + 0.38, mean_v),
                fontsize=7.5, color=color, fontweight="bold",
                va="center", ha="left",
                arrowprops=dict(arrowstyle="-", color=color, lw=0.8))
    ax.annotate(f"★{best_v:.0f}", xy=(pos, best_v),
                xytext=(pos - 0.38, best_v),
                fontsize=7.0, color=color,
                va="center", ha="right",
                arrowprops=dict(arrowstyle="-", color=color, lw=0.8))
    return bp


def _smart_ylim(all_vals, ph_val, pad_frac=0.18):
    """
    Compute a tight but readable y-axis range.
    Minimum span = 15% of the median value, so flat data still shows a box.
    """
    if not all_vals:
        return (0, 1)
    arr = np.array(all_vals, dtype=float)
    lo, hi = float(np.min(arr)), float(np.max(arr))
    if ph_val is not None and ph_val > 0:
        hi = max(hi, ph_val)
    # minimum span: 15% of median value (so e.g. median=10 → min span=1.5)
    min_span = max(float(np.median(arr)) * 0.15, 1.0)
    span = max(hi - lo, min_span)
    centre = (lo + hi) / 2
    pad = span * pad_frac
    return (centre - span / 2 - pad, centre + span / 2 + pad * 1.8)


def _single_ax(ax, case, mh_runs, ma_runs, ph_vals, x_label,
               arpd_mh=None, arpd_ma=None):
    """Render one subplot."""
    mh_data = mh_runs.get(case, [])
    ma_data = ma_runs.get(case, [])
    ph_val  = ph_vals.get(case)

    has_mh = bool(mh_data) and max(mh_data) > 0
    has_ma = bool(ma_data) and max(ma_data) > 0

    # background
    ax.set_facecolor("#F8F9FA")
    ax.grid(axis="y", color="white", linewidth=1.2, zorder=1)
    ax.set_axisbelow(True)
    for spine in ax.spines.values():
        spine.set_visible(False)

    if has_mh:
        _draw_box(ax, 1, mh_data, COLOR_MH)
    if has_ma:
        _draw_box(ax, 2, ma_data, COLOR_MEMETIC)

    # y-axis range (no PH line in plot)
    all_vals = list(mh_data or []) + list(ma_data or [])
    ylo, yhi = _smart_ylim(all_vals, None)
    ax.set_ylim(ylo, yhi)

    # ARPD badges — small colored labels at the bottom of each column
    _badge_y = ylo + (yhi - ylo) * 0.03
    _badge_kw = dict(ha="center", va="bottom", fontsize=7.2,
                     fontweight="bold", zorder=8)
    if arpd_mh is not None and not np.isnan(arpd_mh):
        ax.text(1, _badge_y, f"ARPD: {arpd_mh:.2f}%", color=COLOR_MH,
                bbox=dict(boxstyle="round,pad=0.2", fc="white",
                          ec=COLOR_MH, alpha=0.88, lw=0.9),
                **_badge_kw)
    if arpd_ma is not None and not np.isnan(arpd_ma):
        ax.text(2, _badge_y, f"ARPD: {arpd_ma:.2f}%", color=COLOR_MEMETIC,
                bbox=dict(boxstyle="round,pad=0.2", fc="white",
                          ec=COLOR_MEMETIC, alpha=0.88, lw=0.9),
                **_badge_kw)

    # ticks & labels
    ax.set_xticks([1, 2])
    ax.set_xticklabels(["MH", "Memetic"], fontsize=9, fontweight="bold")
    ax.tick_params(axis="x", length=0, pad=6)
    ax.tick_params(axis="y", labelsize=8, color="#888888")
    ax.set_xlim(0.2, 2.8)

    # N/A notice when data absent
    if not has_ma and not has_mh:
        ax.text(1.5, (ylo + yhi) / 2, "No data", ha="center",
                va="center", fontsize=9, color="#AAAAAA", style="italic")
    elif not has_ma:
        ax.text(2, (ylo + yhi) / 2, "pending", ha="center",
                va="center", fontsize=8, color="#AAAAAA", style="italic")

    ax.set_title(x_label, fontsize=10, fontweight="bold",
                 pad=8, color="#1A1A2E")
    ax.set_ylabel("Fitness (makespan)", fontsize=8, color="#555555")


def make_boxplot_figure(
    cases: list[str],
    mh_runs: dict[str, list[float]],
    ma_runs: dict[str, list[float]],
    ph_vals: dict[str, float],
    title: str,
    out_path: Path,
    x_labels: list[str] | None = None,
    ncols: int = 4,
    arpd_vals: dict[str, tuple] | None = None,
):
    n = len(cases)
    nrows = (n + ncols - 1) // ncols   # ceil division

    fig_w = ncols * 3.8
    fig_h = nrows * 4.6

    fig, axes_grid = plt.subplots(
        nrows, ncols,
        figsize=(fig_w, fig_h),
        sharey=False,
        squeeze=False,
    )
    axes_flat = axes_grid.flatten()

    labels = x_labels if x_labels else cases

    for idx, (case, lbl) in enumerate(zip(cases, labels)):
        amh, ama = (arpd_vals or {}).get(case, (None, None))
        _single_ax(axes_flat[idx], case, mh_runs, ma_runs, ph_vals, lbl,
                   arpd_mh=amh, arpd_ma=ama)

    # hide unused axes
    for idx in range(n, len(axes_flat)):
        axes_flat[idx].set_visible(False)

    patch_mh  = mpatches.Patch(facecolor=COLOR_MH,      label="MH",
                                edgecolor="white", alpha=FACE_ALPHA)
    patch_ma  = mpatches.Patch(facecolor=COLOR_MEMETIC,  label="Memetic",
                                edgecolor="white", alpha=FACE_ALPHA)
    fig.legend(handles=[patch_mh, patch_ma],
               loc="lower center", ncol=3, fontsize=9,
               bbox_to_anchor=(0.5, 0.01),
               frameon=True, framealpha=0.9,
               edgecolor="#CCCCCC")

    fig.patch.set_facecolor("white")
    fig.suptitle(title, fontsize=13, fontweight="bold",
                 color="#1A1A2E", y=1.01)
    plt.tight_layout(rect=[0, 0.06, 1, 0.98])
    fig.savefig(out_path, dpi=180, bbox_inches="tight",
                facecolor="white")
    plt.close(fig)
    print(f"  ✓  Saved: {out_path}")


# ============================================================================
# EXCEL TEST RESULTS
# ============================================================================

HDR  = "1F4E79"
INFO = "2E75B6"
THIN = Side(style="thin", color="AAAAAA")
BDR  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOOD = "C6EFCE"
BAD  = "FFCCCC"
NEU  = "FFEB9C"


def _h(ws, r, c, v, bg=HDR):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border    = BDR
    return cell


def _c(ws, r, c, v, bg="FFFFFF", bold=False, fmt=None, align="center"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(bold=bold, name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = BDR
    if fmt:
        cell.number_format = fmt
    return cell


def save_test_excel(
    test_rows: list[dict],
    out_path: Path,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Statistical Tests"
    ws.sheet_view.showGridLines = False

    # Headers
    _h(ws, 1, 1, "Instance", bg="1F3864")
    _h(ws, 1, 2, "Group",    bg="1F3864")
    for ci, h in enumerate(["MH mean", "MH std", "MH best"], 3):
        _h(ws, 1, ci, h, bg="1F4E79")
    for ci, h in enumerate(["MA mean", "MA std", "MA best"], 6):
        _h(ws, 1, ci, h, bg="375623")
    _h(ws, 1, 9,  "Δ mean (MA−MH)", bg="1F3864")
    # ARPD columns (cols 10-11)
    _h(ws, 1, 10, "ARPD MH (%)",      bg="1F4E79")
    _h(ws, 1, 11, "ARPD Memetic (%)", bg="375623")
    _h(ws, 1, 12, "ΔARPD (MA−MH)",    bg="1F3864")
    for ci, h in enumerate(["Wilcoxon stat", "Wilcoxon p", "Sig. (W)"], 13):
        _h(ws, 1, ci, h, bg="7030A0")
    for ci, h in enumerate(["Paired t stat", "Paired t p", "Sig. (t)"], 16):
        _h(ws, 1, ci, h, bg="843C0C")
    _h(ws, 1, 19, "Conclusion", bg="1F3864")

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = ws.cell(row=2, column=1)

    def _sig_label(p):
        if np.isnan(p): return "N/A"
        if p < 0.001:   return "*** p<0.001"
        if p < 0.01:    return "**  p<0.01"
        if p < 0.05:    return "*   p<0.05"
        return "n.s."

    for ri, row in enumerate(test_rows, 2):
        bg = "FFFFFF" if ri % 2 == 0 else "F5F5F5"
        delta     = row["ma_mean"] - row["mh_mean"]
        arpd_mh   = row.get("arpd_mh", np.nan)
        arpd_ma   = row.get("arpd_ma", np.nan)
        delta_arpd = (arpd_ma - arpd_mh) if not (np.isnan(arpd_mh) or np.isnan(arpd_ma)) else np.nan
        conc      = row["conclusion"]
        conc_bg   = (GOOD if "Memetic better" in conc
                     else BAD if "MH better" in conc else NEU)

        vals = [
            row["instance"], row["group"],
            row["mh_mean"], row["mh_std"], row["mh_best"],
            row["ma_mean"], row["ma_std"], row["ma_best"],
            delta,
            arpd_mh, arpd_ma, delta_arpd,
            row.get("w_stat"), row.get("w_p"), _sig_label(row.get("w_p", float("nan"))),
            row.get("t_stat"), row.get("t_p"), _sig_label(row.get("t_p", float("nan"))),
            conc,
        ]
        fmts = [None, None,
                "0.0000", "0.0000", "0.0000",
                "0.0000", "0.0000", "0.0000",
                "+0.0000;-0.0000",
                "0.00\"%\"", "0.00\"%\"", "+0.00\"%\";-0.00\"%\"",
                "0.0000", "0.0000", None,
                "0.0000", "0.0000", None,
                None]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = _c(ws, ri, ci, v, bg=bg, fmt=fmt,
                      align="left" if ci in (1, 2, 15, 18, 19) else "center")
            if ci == 9:
                cell.fill = PatternFill("solid", start_color=GOOD if delta < 0 else BAD)
            if ci == 10:  # ARPD MH
                cell.fill = PatternFill("solid", start_color="DDEEFF")
            if ci == 11:  # ARPD Memetic
                cell.fill = PatternFill("solid", start_color="DDFFDD")
            if ci == 12:  # ΔARPD
                if not np.isnan(delta_arpd):
                    cell.fill = PatternFill("solid", start_color=GOOD if delta_arpd < 0 else BAD)
            if ci == 15:  # Wilcoxon sig label
                w_p = row.get("w_p", float("nan"))
                cell.fill = PatternFill("solid",
                    start_color=GOOD if (not np.isnan(w_p) and w_p < 0.05) else NEU)
            if ci == 18:  # t-test sig label
                t_p = row.get("t_p", float("nan"))
                cell.fill = PatternFill("solid",
                    start_color=GOOD if (not np.isnan(t_p) and t_p < 0.05) else NEU)
            if ci == 19:
                cell.fill = PatternFill("solid", start_color=conc_bg)
        ws.row_dimensions[ri].height = 16

    widths = [22, 8, 9, 8, 8, 9, 8, 8, 14, 12, 14, 14, 13, 11, 13, 13, 11, 13, 20]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(out_path)
    print(f"  ✓  Saved: {out_path}")


# ============================================================================
# MAIN
# ============================================================================

# ── Key mapping helpers ──────────────────────────────────────────────────────
# Sheet names differ slightly between files; we normalise by ops count.

def _normalise_key(raw: str) -> str:
    """Strip surrounding spaces/underscores and collapse whitespace."""
    return " ".join(raw.replace("__", " ").replace("_", " ").split()).strip()


def _ops_from_key(key: str) -> int | None:
    """Extract ops number from labels like '6M140 n=60' or '6M140-n60 (6M-60ops)'."""
    import re
    m = re.search(r"(\d+)\s*ops", key, re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(r"n[=\-]?(\d+)", key, re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)\s*$", key)
    if m:
        return int(m.group(1))
    return None


def _build_ops_map(d: dict[str, list]) -> dict[int, list]:
    """Map ops-count → run list, for cross-file matching."""
    out = {}
    for k, v in d.items():
        n = _ops_from_key(k)
        if n is not None:
            out[n] = v
    return out


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── Load available files ─────────────────────────────────────────────
    files = {
        "mh_base":   Path(MH_BASE_XLSX),
        "mh_t8":     Path(MH_TABLE8_XLSX),
        "ma_base":   Path(MEMETIC_BASE_XLSX),
        "ma_t8":     Path(MEMETIC_T8_XLSX),
    }
    loaded: dict[str, dict] = {}
    ph_base: dict[str, float] = {}
    ph_t8:   dict[str, float] = {}

    for key, path in files.items():
        if path.exists():
            print(f"Reading {path} ...")
            loaded[key] = _read_run_fitness(path)
            if key == "mh_base":
                ph_base = _read_summary_ph(path)
            elif key == "mh_t8":
                ph_t8 = _read_summary_ph(path)
        else:
            print(f"⚠️  Not found (skipping): {path}")
            loaded[key] = {}

    # ── Normalise keys ───────────────────────────────────────────────────
    mh_base_data = {_normalise_key(k): v for k, v in loaded["mh_base"].items()}
    mh_t8_data   = {_normalise_key(k): v for k, v in loaded["mh_t8"].items()}
    ma_base_data = {_normalise_key(k): v for k, v in loaded["ma_base"].items()}
    ma_t8_data   = {_normalise_key(k): v for k, v in loaded["ma_t8"].items()}

    ph_base_norm = {_normalise_key(k): v for k, v in ph_base.items()}
    ph_t8_norm   = {_normalise_key(k): v for k, v in ph_t8.items()}

    # For cross-file matching, build ops-indexed maps
    mh_base_ops = _build_ops_map(mh_base_data)
    mh_t8_ops   = _build_ops_map(mh_t8_data)
    ma_base_ops = _build_ops_map(ma_base_data)
    ma_t8_ops   = _build_ops_map(ma_t8_data)
    ph_base_ops = {_ops_from_key(k): v for k, v in ph_base_norm.items()
                   if _ops_from_key(k) is not None}
    ph_t8_ops   = {_ops_from_key(k): v for k, v in ph_t8_norm.items()
                   if _ops_from_key(k) is not None}

    # ── BASE CASES (Table 14) ────────────────────────────────────────────
    base_ops_order = [38, 46, 140, 163]
    base_labels    = ["2M38 (38 ops)", "2M46 (46 ops)",
                      "6M140 (140 ops)", "6M163 (163 ops)"]

    base_test_rows = []
    arpd_for_plot_base: dict[str, tuple] = {}
    for n_ops, label in zip(base_ops_order, base_labels):
        mh_r = mh_base_ops.get(n_ops, [])
        ma_r = ma_base_ops.get(n_ops, [])
        if not mh_r and not ma_r:
            continue

        all_r = list(mh_r) + list(ma_r)
        bks   = float(np.min(all_r)) if all_r else np.nan
        amh   = _compute_arpd(mh_r, bks)
        ama   = _compute_arpd(ma_r, bks)
        arpd_for_plot_base[str(n_ops)] = (amh, ama)

        test_res = _run_test(mh_r, ma_r) if (mh_r and ma_r) else \
                   {"w_stat": np.nan, "w_p": np.nan, "t_stat": np.nan, "t_p": np.nan, "conclusion": "N/A"}
        base_test_rows.append({
            "instance": label, "group": "Base",
            "mh_mean": float(np.mean(mh_r)) if mh_r else np.nan,
            "mh_std":  float(np.std(mh_r))  if mh_r else np.nan,
            "mh_best": float(np.min(mh_r))  if mh_r else np.nan,
            "ma_mean": float(np.mean(ma_r)) if ma_r else np.nan,
            "ma_std":  float(np.std(ma_r))  if ma_r else np.nan,
            "ma_best": float(np.min(ma_r))  if ma_r else np.nan,
            "arpd_mh": amh,
            "arpd_ma": ama,
            **test_res,
        })

    # Box plot – base cases
    print("\nGenerating base-cases box plot...")
    ph_for_plot_base = {str(n): ph_base_ops.get(n) for n in base_ops_order}
    mh_for_plot_base = {str(n): mh_base_ops.get(n, []) for n in base_ops_order}
    ma_for_plot_base = {str(n): ma_base_ops.get(n, []) for n in base_ops_order}
    make_boxplot_figure(
        cases=[str(n) for n in base_ops_order],
        mh_runs=mh_for_plot_base,
        ma_runs=ma_for_plot_base,
        ph_vals=ph_for_plot_base,
        title="MH vs Memetic – Base Cases (Table 14)",
        out_path=OUTPUT_DIR / "compare_base_cases.png",
        x_labels=base_labels,
        ncols=4,
        arpd_vals=arpd_for_plot_base,
    )

    # ── TABLE 8 (sub-instances of 6M140) ────────────────────────────────
    t8_ops_order = [15, 25, 30, 60, 90, 120, 140]
    t8_labels    = [f"n={n}" for n in t8_ops_order]

    t8_test_rows = []
    arpd_for_plot_t8: dict[str, tuple] = {}
    for n_ops, label in zip(t8_ops_order, t8_labels):
        mh_r = mh_t8_ops.get(n_ops, [])
        ma_r = ma_t8_ops.get(n_ops, [])
        if not mh_r and not ma_r:
            continue

        all_r = list(mh_r) + list(ma_r)
        bks   = float(np.min(all_r)) if all_r else np.nan
        amh   = _compute_arpd(mh_r, bks)
        ama   = _compute_arpd(ma_r, bks)
        arpd_for_plot_t8[str(n_ops)] = (amh, ama)

        test_res = _run_test(mh_r, ma_r) if (mh_r and ma_r) else \
                   {"w_stat": np.nan, "w_p": np.nan, "t_stat": np.nan, "t_p": np.nan, "conclusion": "N/A"}
        t8_test_rows.append({
            "instance": f"6M140 {label}", "group": "Table 8",
            "mh_mean": float(np.mean(mh_r)) if mh_r else np.nan,
            "mh_std":  float(np.std(mh_r))  if mh_r else np.nan,
            "mh_best": float(np.min(mh_r))  if mh_r else np.nan,
            "ma_mean": float(np.mean(ma_r)) if ma_r else np.nan,
            "ma_std":  float(np.std(ma_r))  if ma_r else np.nan,
            "ma_best": float(np.min(ma_r))  if ma_r else np.nan,
            "arpd_mh": amh,
            "arpd_ma": ama,
            **test_res,
        })

    # Box plot – Table 8
    print("Generating Table-8 box plot...")
    ph_for_plot_t8 = {str(n): ph_t8_ops.get(n) for n in t8_ops_order}
    mh_for_plot_t8 = {str(n): mh_t8_ops.get(n, []) for n in t8_ops_order}
    ma_for_plot_t8 = {str(n): ma_t8_ops.get(n, []) for n in t8_ops_order}
    make_boxplot_figure(
        cases=[str(n) for n in t8_ops_order],
        mh_runs=mh_for_plot_t8,
        ma_runs=ma_for_plot_t8,
        ph_vals=ph_for_plot_t8,
        title="MH vs Memetic – Table 8 sub-instances of 6M140\n(sorted by release time, first n ops)",
        out_path=OUTPUT_DIR / "compare_table8.png",
        x_labels=t8_labels,
        ncols=4,
        arpd_vals=arpd_for_plot_t8,
    )

    # ── Statistical tests Excel ──────────────────────────────────────────
    print("\nSaving statistical tests to Excel...")
    all_rows = base_test_rows + t8_test_rows
    if all_rows:
        save_test_excel(all_rows, OUTPUT_DIR / "statistical_tests.xlsx")

    # ── Console summary ──────────────────────────────────────────────────
    print("\n" + "=" * 110)
    print(f"{'Instance':<22} {'Group':<8} {'MH mean':>8} {'MA mean':>8} "
          f"{'Δ':>8} {'ARPD MH':>9} {'ARPD MA':>9} {'ΔARPD':>9} "
          f"{'W p':>8} {'t p':>8} {'Conclusion'}")
    print("-" * 110)
    for r in all_rows:
        delta      = r["ma_mean"] - r["mh_mean"]
        arpd_mh    = r.get("arpd_mh", float("nan"))
        arpd_ma    = r.get("arpd_ma", float("nan"))
        delta_arpd = (arpd_ma - arpd_mh) if not (np.isnan(arpd_mh) or np.isnan(arpd_ma)) else float("nan")
        w_p        = r.get("w_p", float("nan"))
        t_p        = r.get("t_p", float("nan"))

        def _fmt_f(v, spec):
            return f"{v:{spec}}" if not np.isnan(v) else "N/A"

        print(f"{r['instance']:<22} {r['group']:<8} "
              f"{r['mh_mean']:>8.3f} {r['ma_mean']:>8.3f} "
              f"{delta:>+8.3f} "
              f"{(_fmt_f(arpd_mh, '.2f')+'%'):>9} "
              f"{(_fmt_f(arpd_ma, '.2f')+'%'):>9} "
              f"{(_fmt_f(delta_arpd, '+.2f')+'%' if not np.isnan(delta_arpd) else 'N/A'):>9} "
              f"{_fmt_f(w_p, '.4f'):>8} "
              f"{_fmt_f(t_p, '.4f'):>8}  "
              f"{r['conclusion']}")
    print("=" * 110)
    print("\nDone.")


if __name__ == "__main__":
    main()
