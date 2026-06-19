"""
export.py – Save experiment results to a formatted Excel file.

Saves a .xlsx file with one Summary sheet and one detail sheet per instance.
Contains two entry points:
  - save_memetic_excel: for Memetic Algorithm standalone results (matches reference file format)
  - save_mh_only_excel: for MH results
"""
# ── Imports ──────────────────────────────────────────────────────────────
from __future__ import annotations
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Shared helpers ────────────────────────────────────────────────────────

HDR   = "1F4E79"
INFO  = "2E75B6"
BEST  = "C6EFCE"
WORST = "FFCCCC"
THIN  = Side(style="thin", color="AAAAAA")
BDR   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr(ws, r, c, txt, bg=HDR, bold=True, white=True):
    cell = ws.cell(row=r, column=c, value=txt)
    cell.font      = Font(bold=bold, color="FFFFFF" if white else "000000",
                          name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BDR
    return cell


def _cel(ws, r, c, val, bg="FFFFFF", bold=False, fmt=None, align="center"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(bold=bold, name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = BDR
    if fmt:
        cell.number_format = fmt
    return cell


def _write_meta_block(ws, start_row, n_cols, title, run_name, timestamp, params):
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row,   end_column=n_cols)
    t = ws.cell(row=start_row, column=1, value=title)
    t.font      = Font(bold=True, size=13, name="Arial", color="FFFFFF")
    t.fill      = PatternFill("solid", start_color=HDR)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 22
    r = start_row + 1

    for label, value, bg in [
        ("Experiment name", run_name,  "EBF3FB"),
        ("Date / time",     timestamp, "FFFFFF"),
        ("Parameters",      "   |   ".join(f"{k} = {v}" for k, v in params.items()), "FFFFFF"),
    ]:
        _hdr(ws, r, 1, label, bg=INFO)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=n_cols)
        c = ws.cell(row=r, column=2, value=value)
        c.font      = Font(bold=(label == "Experiment name"), name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = BDR
        ws.row_dimensions[r].height = 16
        r += 1

    ws.row_dimensions[r].height = 6
    return r + 1


# ── MH vs MH+LS ───────────────────────────────────────────────────────────

def save_mhls_excel(all_results: List[Dict],
                    output_path: str = "results_mhls.xlsx",
                    meta: Optional[Dict] = None):
    """Write a formatted Excel workbook comparing MH vs MH+LS.
    Expects result dicts with keys:
        label, n_ops, n_machines, n_tools, capacity, ph_obj,
        mh_mean, mh_std, mh_best, mh_arpd,
        mhls_mean, mhls_std, mhls_best, mhls_arpd,
        and optionally: mh_run_details = list of (fitness, time, gens) tuples,
                      mhls_run_details = list of (fitness, time, gens) tuples.
    """
    if meta is None:
        meta = {}
    run_name  = meta.get("run_name",  "—")
    timestamp = meta.get("timestamp", "—")
    params    = meta.get("params",    {})

    COL_MH   = "2E75B6"   # blue  – MH
    COL_MHLS = "375623"   # green – MH+LS

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    N_SUMMARY_COLS = 16
    next_row = _write_meta_block(ws, 1, N_SUMMARY_COLS,
                                 "MH vs MH+LS Results – Local Search post-processing",
                                 run_name, timestamp, params)

    # ── Summary headers ──
    headers = ["Instance", "Ops", "Machines", "Tools", "Capacity", "PH obj",
               "MH mean", "MH std", "MH best", "MH ARPD(%)",
               "MH+LS mean", "MH+LS std", "MH+LS best", "MH+LS ARPD(%)",
               "Δ mean", "Δ ARPD(%)"]
    col_bgs = [HDR]*6 + [COL_MH]*4 + [COL_MHLS]*4 + [HDR]*2
    for ci, (h, bg) in enumerate(zip(headers, col_bgs), 1):
        _hdr(ws, next_row, ci, h, bg=bg)
    ws.row_dimensions[next_row].height = 30
    ws.freeze_panes = ws.cell(row=next_row + 1, column=1)
    next_row += 1

    for ri, res in enumerate(all_results):
        bg = "FFFFFF" if ri % 2 == 0 else "F5F5F5"
        delta_mean = res["mhls_mean"] - res["mh_mean"]
        delta_arpd = res["mhls_arpd"] - res["mh_arpd"]
        vals = [res["label"], res["n_ops"], res["n_machines"], res["n_tools"],
                res["capacity"], res["ph_obj"],
                res["mh_mean"],   res["mh_std"],   res["mh_best"],   res["mh_arpd"],
                res["mhls_mean"], res["mhls_std"], res["mhls_best"], res["mhls_arpd"],
                delta_mean, delta_arpd]
        fmts = [None, "0", "0", "0", "0", "0.00",
                "0.00", "0.00", "0.00", "0.00",
                "0.00", "0.00", "0.00", "0.00",
                "+0.00;-0.00", "+0.00;-0.00"]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            _cel(ws, next_row, ci, v, bg=bg, fmt=fmt,
                 align="left" if ci == 1 else "center")
        # Color the Δ columns: green if MH+LS improved (negative delta = lower cost)
        for col, val in [(15, delta_mean), (16, delta_arpd)]:
            ws.cell(row=next_row, column=col).fill = PatternFill(
                "solid", start_color=BEST if val < 0 else WORST)
        ws.row_dimensions[next_row].height = 16
        next_row += 1

    for ci, w in enumerate([22, 6, 9, 7, 9, 8,
                             8, 7, 8, 10,
                             10, 7, 10, 12,
                             10, 10], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Detail sheets ──
    for res in all_results:
        sname = res["label"].replace(" ", "_")[:31]
        ws2 = wb.create_sheet(title=sname)
        ws2.sheet_view.showGridLines = False

        N_DETAIL_COLS = 7
        next_row2 = _write_meta_block(ws2, 1, N_DETAIL_COLS,
                                      "MH vs MH+LS – Run detail",
                                      run_name, timestamp, params)

        # Instance info
        for i, (k, v) in enumerate(
                [("Operations",   res["n_ops"]),
                 ("Machines",     res["n_machines"]),
                 ("Tool sets",    res["n_tools"]),
                 ("Capacity",     res["capacity"]),
                 ("PH objective", res["ph_obj"])]):
            _hdr(ws2, next_row2 + i, 1, k, bg=INFO)
            _cel(ws2, next_row2 + i, 2, v,
                 fmt="0.00" if isinstance(v, float) else "0")
        next_row2 += 6

        # Run-by-run table
        _hdr(ws2, next_row2, 1, "Run",   bg=INFO)
        for ci, h in enumerate(["Fitness", "Time(s)", "Gens"], 2):
            _hdr(ws2, next_row2, ci, f"MH – {h}",    bg=COL_MH)
        for ci, h in enumerate(["Fitness", "Time(s)", "Gens"], 5):
            _hdr(ws2, next_row2, ci, f"MH+LS – {h}", bg=COL_MHLS)
        next_row2 += 1

        for i, ((mh_f, mh_t, mh_g), (ls_f, ls_t, ls_g)) in enumerate(
                zip(res["mh_run_details"], res["mhls_run_details"])):
            bg = "F0F7FF" if i % 2 == 0 else "FFFFFF"
            _cel(ws2, next_row2, 1, i + 1,  bg=bg, fmt="0")
            _cel(ws2, next_row2, 2, mh_f,   bg=bg, fmt="0.0000")
            _cel(ws2, next_row2, 3, mh_t,   bg=bg, fmt="0.0")
            _cel(ws2, next_row2, 4, mh_g,   bg=bg, fmt="0")
            _cel(ws2, next_row2, 5, ls_f,   bg=bg, fmt="0.0000")
            _cel(ws2, next_row2, 6, ls_t,   bg=bg, fmt="0.0")
            _cel(ws2, next_row2, 7, ls_g,   bg=bg, fmt="0")
            # Highlight best fitness between MH and MH+LS
            for col, val, ref in [(2, mh_f, ls_f), (5, ls_f, mh_f)]:
                ws2.cell(row=next_row2, column=col).fill = PatternFill(
                    "solid", start_color=BEST  if val < ref else
                                        (WORST if val > ref else "FFFFFF"))
            next_row2 += 1

        next_row2 += 1

        # Statistics comparison block
        _hdr(ws2, next_row2, 1, "Statistic", bg=HDR)
        _hdr(ws2, next_row2, 2, "MH",        bg=COL_MH)
        _hdr(ws2, next_row2, 3, "MH+LS",     bg=COL_MHLS)
        _hdr(ws2, next_row2, 4, "Δ",         bg="7030A0")
        next_row2 += 1
        for lbl, mv, lv in zip(
                ["Mean", "Std", "Best", "Avg time(s)", "ARPD(%)"],
                [res["mh_mean"],   res["mh_std"],   res["mh_best"],
                 res["mh_time"],   res["mh_arpd"]],
                [res["mhls_mean"], res["mhls_std"], res["mhls_best"],
                 res["mhls_time"], res["mhls_arpd"]]):
            _cel(ws2, next_row2, 1, lbl, bold=True, align="left")
            _cel(ws2, next_row2, 2, mv,  fmt="0.0000")
            _cel(ws2, next_row2, 3, lv,  fmt="0.0000")
            d  = lv - mv
            dc = _cel(ws2, next_row2, 4, d, fmt="+0.0000;-0.0000")
            dc.fill = PatternFill("solid", start_color=BEST if d < 0 else WORST)
            next_row2 += 1

        for ci, w in enumerate([12, 12, 10, 8, 12, 10, 8], 1):
            ws2.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)
    print(f"\n✓  Results saved to: {output_path}")

# ── Memetic Algorithm standalone ──────────────────────────────────────────
def save_memetic_excel(all_results: List[Dict],
                       output_path: str = "results_memetic.xlsx",
                       meta: Optional[Dict] = None):
    """Write a formatted Excel workbook for Memetic Algorithm standalone results.

    Matches the exact format of results_mh_base.xlsx (MH Standalone Results).

    Expects result dicts with keys:
        label, n_ops, n_machines, n_tools, capacity,
        ph_mean, mean, std, best, worst, arpd_mean, total_time, runs,
        and optionally: run_details = list of (fitness, time, gens) tuples.
    """
    if meta is None:
        meta = {}
    run_name  = meta.get("run_name",  "—")
    timestamp = meta.get("timestamp", "—")
    params    = meta.get("params",    {})

    COL_MA = "2E75B6"   # blue – same as MH in reference file

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    N_SUMMARY_COLS = 13
    next_row = _write_meta_block(ws, 1, N_SUMMARY_COLS,
                                 "MH Standalone Results",
                                 run_name, timestamp, params)

    # Summary headers – mirrors reference file exactly
    headers = ["Instance", "Ops", "Machines", "Tools", "Capacity", "Stop criterion",
               "PH obj (ref)",
               "MH mean", "MH std", "MH best", "MH avg time(s)", "MH avg gens",
               "MH ARPD(%)"]
    col_bgs = [HDR] * 7 + [COL_MA] * 5 + [HDR]
    for ci, (h, bg) in enumerate(zip(headers, col_bgs), 1):
        _hdr(ws, next_row, ci, h, bg=bg)
    ws.row_dimensions[next_row].height = 30
    ws.freeze_panes = ws.cell(row=next_row + 1, column=1)
    next_row += 1

    for ri, res in enumerate(all_results):
        bg = "FFFFFF" if ri % 2 == 0 else "F5F5F5"
        stop_str = meta.get("stop_str", f"evals={meta.get('params', {}).get('evals', '?')}")
        avg_time = res.get("avg_time", res.get("total_time", 0) / max(len(res.get("runs", [1])), 1))
        avg_gens = res.get("avg_gens", res.get("mean_gens", 0))
        vals = [
            res["label"], res["n_ops"],
            res.get("n_machines", "—"), res.get("n_tools", "—"),
            res.get("capacity", "—"), stop_str,
            res["ph_mean"],
            res["mean"], res["std"], res["best"],
            avg_time, avg_gens,
            res["arpd_mean"],
        ]
        fmts = [None, "0", "0", "0", "0", None,
                "0.0",
                "0.0", "0.0000", "0.0", "0.000000000000000", "0.0",
                "0.00"]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            _cel(ws, next_row, ci, v, bg=bg, fmt=fmt,
                 align="left" if ci == 1 else "center")
        # Colour ARPD: green if MA beat PH (negative ARPD = lower cost)
        ws.cell(row=next_row, column=13).fill = PatternFill(
            "solid", start_color=BEST if res["arpd_mean"] < 0 else WORST)
        ws.row_dimensions[next_row].height = 16
        next_row += 1

    for ci, w in enumerate([22, 6, 9, 7, 9, 14, 12,
                             8, 7, 8, 16, 11, 11], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Detail sheets ──
    for res in all_results:
        sname = res["label"].replace(" ", "_")[:31]
        ws2 = wb.create_sheet(title=sname)
        ws2.sheet_view.showGridLines = False

        N_DETAIL_COLS = 4
        next_row2 = _write_meta_block(ws2, 1, N_DETAIL_COLS,
                                      "MH – Run detail",
                                      run_name, timestamp, params)

        stop_str = meta.get("stop_str", f"evals={meta.get('params', {}).get('evals', '?')}")
        info_rows = [
            ("Operations",   res["n_ops"]),
            ("Machines",     res.get("n_machines", "—")),
            ("Tool sets",    res.get("n_tools",    "—")),
            ("Capacity",     res.get("capacity",   "—")),
            ("Stop",         stop_str),
            ("PH objective", res["ph_mean"]),
        ]
        for i, (k, v) in enumerate(info_rows):
            _hdr(ws2, next_row2 + i, 1, k, bg=INFO)
            _cel(ws2, next_row2 + i, 2, v,
                 fmt="0.00" if isinstance(v, float) else None)
        next_row2 += len(info_rows) + 1

        # Run-by-run table header
        _hdr(ws2, next_row2, 1, "Run",      bg=INFO)
        _hdr(ws2, next_row2, 2, "Fitness",  bg=COL_MA)
        _hdr(ws2, next_row2, 3, "Time(s)",  bg=COL_MA)
        _hdr(ws2, next_row2, 4, "Gens",     bg=COL_MA)
        next_row2 += 1

        run_details = res.get("run_details", [])
        best_f = min((f for f, _, _ in run_details), default=None) if run_details else None

        for i, (f, t, g) in enumerate(run_details):
            bg = "F0F7FF" if i % 2 == 0 else "FFFFFF"
            _cel(ws2, next_row2, 1, i + 1, bg=bg, fmt="0")
            fc = _cel(ws2, next_row2, 2, f, bg=bg, fmt="0")
            _cel(ws2, next_row2, 3, t, bg=bg, fmt="0.000000000000000")
            _cel(ws2, next_row2, 4, g, bg=bg, fmt="0")
            if best_f is not None and f == best_f:
                fc.fill = PatternFill("solid", start_color=BEST)
            next_row2 += 1

        next_row2 += 1

        # Statistics block
        _hdr(ws2, next_row2, 1, "Statistic", bg=HDR)
        _hdr(ws2, next_row2, 2, "Value",     bg=COL_MA)
        next_row2 += 1

        avg_time = res.get("avg_time", res.get("total_time", 0) / max(len(res.get("runs", [1])), 1))
        avg_gens = res.get("avg_gens", res.get("mean_gens", 0))
        for lbl, val, fmt in [
            ("Mean",           res["mean"],      "0"),
            ("Std",            res["std"],        "0.0000000000000000"),
            ("Best",           res["best"],       "0"),
            ("Avg time(s)",    avg_time,          "0.000000000000000"),
            ("Avg gens",       avg_gens,          "0.0"),
            ("ARPD vs PH(%)",  res["arpd_mean"],  "0.00"),
        ]:
            _cel(ws2, next_row2, 1, lbl, bold=True, align="left")
            _cel(ws2, next_row2, 2, val, fmt=fmt)
            next_row2 += 1

        for ci, w in enumerate([14, 12, 20, 8], 1):
            ws2.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)
    print(f"\n✓  Results saved to: {output_path}")

# ── MH standalone (run_mh.py entry point) ────────────────────────────────
 
def save_mh_only_excel(all_results,
                       output_path: str = "results_mh_only.xlsx",
                       meta=None):
    """Write a formatted Excel workbook for MH-only standalone results.
 
    Expects result dicts produced by run_mh.py's run_instance(), with keys:
        label, n_ops, n_machines, n_tools, capacity,
        ph_obj, ph_std, stop_criterion,
        mh_mean, mh_std, mh_best, mh_time, mh_gens, mh_arpd,
        mh_run_details  (list of (fitness, time, gens) tuples)
    """
    if meta is None:
        meta = {}
    run_name  = meta.get("run_name",  "—")
    timestamp = meta.get("timestamp", "—")
    params    = meta.get("params",    {})
 
    COL_MH = "2E75B6"   # blue
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
 
    N_SUMMARY_COLS = 13
    next_row = _write_meta_block(ws, 1, N_SUMMARY_COLS,
                                 "MH Standalone Results",
                                 run_name, timestamp, params)
 
    headers = ["Instance", "Ops", "Machines", "Tools", "Capacity", "Stop criterion",
               "PH obj (ref)",
               "MH mean", "MH std", "MH best", "MH avg time(s)", "MH avg gens",
               "MH ARPD(%)"]
    col_bgs = [HDR] * 7 + [COL_MH] * 5 + [HDR]
    for ci, (h, bg) in enumerate(zip(headers, col_bgs), 1):
        _hdr(ws, next_row, ci, h, bg=bg)
    ws.row_dimensions[next_row].height = 30
    ws.freeze_panes = ws.cell(row=next_row + 1, column=1)
    next_row += 1
 
    for ri, res in enumerate(all_results):
        bg = "FFFFFF" if ri % 2 == 0 else "F5F5F5"
        stop_str = res.get("stop_criterion", meta.get("stop_str", "—"))
        vals = [
            res["label"], res["n_ops"],
            res.get("n_machines", "—"), res.get("n_tools", "—"),
            res.get("capacity",   "—"), stop_str,
            res["ph_obj"],
            res["mh_mean"], res["mh_std"], res["mh_best"],
            res["mh_time"], res["mh_gens"],
            res["mh_arpd"],
        ]
        fmts = [None, "0", "0", "0", "0", None,
                "0.00",
                "0.0000", "0.0000", "0.0000", "0.0", "0.0",
                "0.00"]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            _cel(ws, next_row, ci, v, bg=bg, fmt=fmt,
                 align="left" if ci == 1 else "center")
        ws.cell(row=next_row, column=13).fill = PatternFill(
            "solid", start_color=BEST if res["mh_arpd"] < 0 else WORST)
        ws.row_dimensions[next_row].height = 16
        next_row += 1
 
    for ci, w in enumerate([22, 6, 9, 7, 9, 14, 12,
                             8, 7, 8, 14, 10, 11], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
 
    # ── Detail sheets ──
    for res in all_results:
        sname = res["label"].replace(" ", "_")[:31]
        ws2 = wb.create_sheet(title=sname)
        ws2.sheet_view.showGridLines = False
 
        N_DETAIL_COLS = 4
        next_row2 = _write_meta_block(ws2, 1, N_DETAIL_COLS,
                                      "MH – Run detail",
                                      run_name, timestamp, params)
 
        stop_str = res.get("stop_criterion", meta.get("stop_str", "—"))
        info_rows = [
            ("Operations",   res["n_ops"]),
            ("Machines",     res.get("n_machines", "—")),
            ("Tool sets",    res.get("n_tools",    "—")),
            ("Capacity",     res.get("capacity",   "—")),
            ("Stop",         stop_str),
            ("PH objective", res["ph_obj"]),
        ]
        for i, (k, v) in enumerate(info_rows):
            _hdr(ws2, next_row2 + i, 1, k, bg=INFO)
            _cel(ws2, next_row2 + i, 2, v,
                 fmt="0.00" if isinstance(v, float) else None)
        next_row2 += len(info_rows) + 1
 
        _hdr(ws2, next_row2, 1, "Run",      bg=INFO)
        _hdr(ws2, next_row2, 2, "Fitness",  bg=COL_MH)
        _hdr(ws2, next_row2, 3, "Time(s)",  bg=COL_MH)
        _hdr(ws2, next_row2, 4, "Gens",     bg=COL_MH)
        next_row2 += 1
 
        run_details = res.get("mh_run_details", [])
        best_f = min((f for f, _, _ in run_details), default=None) if run_details else None
 
        for i, (f, t, g) in enumerate(run_details):
            bg = "F0F7FF" if i % 2 == 0 else "FFFFFF"
            _cel(ws2, next_row2, 1, i + 1, bg=bg, fmt="0")
            fc = _cel(ws2, next_row2, 2, f,   bg=bg, fmt="0.0000")
            _cel(ws2, next_row2, 3, t,         bg=bg, fmt="0.0")
            _cel(ws2, next_row2, 4, g,         bg=bg, fmt="0")
            if best_f is not None and f == best_f:
                fc.fill = PatternFill("solid", start_color=BEST)
            next_row2 += 1
 
        next_row2 += 1
        _hdr(ws2, next_row2, 1, "Statistic", bg=HDR)
        _hdr(ws2, next_row2, 2, "Value",     bg=COL_MH)
        next_row2 += 1
        for lbl, val, fmt in [
            ("Mean",          res["mh_mean"], "0.0000"),
            ("Std",           res["mh_std"],  "0.0000"),
            ("Best",          res["mh_best"], "0.0000"),
            ("Avg time(s)",   res["mh_time"], "0.0"),
            ("Avg gens",      res["mh_gens"], "0.0"),
            ("ARPD vs PH(%)", res["mh_arpd"], "0.00"),
        ]:
            _cel(ws2, next_row2, 1, lbl, bold=True, align="left")
            _cel(ws2, next_row2, 2, val, fmt=fmt)
            next_row2 += 1
 
        for ci, w in enumerate([14, 12, 10, 8], 1):
            ws2.column_dimensions[get_column_letter(ci)].width = w
 
    wb.save(output_path)
    print(f"\n✓  Results saved to: {output_path}")